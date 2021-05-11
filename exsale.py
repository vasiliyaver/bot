from openpyxl import load_workbook


book = load_workbook('Книга2.xlsx')
sheet_1 = book['Лист1']
stickers_page = book['Вася']
print(book.worksheets)
for i in range(1, 6):
    print(stickers_page.cell(row=i, column=1).value)