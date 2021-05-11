from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import Updater, MessageHandler, Filters, CommandHandler
from openpyxl import load_workbook

TOKEN = '1602888769:AAEZpsCLio1mJvcumA-q1M3bwZSzmNhTaCc'

book = load_workbook('Книга2.xlsx')
sheet_1 = book['Лист1']
stickers_page = book['Вася']

def main():
    updater = Updater(token=TOKEN)  # объект, который ловит сообщения из телеграм
    dispatcher = updater.dispatcher
    handler = MessageHandler(Filters.all, do_echo)  # отфильтровываем сообщения: теперь устройство должно реагировать
    start_handler = CommandHandler('start', do_start)
    help_handler = CommandHandler('help', do_help)
    keyboard_handler = MessageHandler(Filters.text, do_something)
    sticker_handler = MessageHandler(Filters.sticker, do_sticker)

    dispatcher.add_handler(start_handler)
    dispatcher.add_handler(help_handler)
    dispatcher.add_handler(keyboard_handler)
    dispatcher.add_handler(sticker_handler)
    dispatcher.add_handler(handler)

    updater.start_polling()
    updater.idle()

def do_echo(update: Update, context):
    user = update.message.from_user.is_bot
    name = update.message.from_user.first_name
    if user:
        update.message.reply_text(text=f"Ты - бот! Уходи отсюда!!!")
    else:
        update.message.reply_text(text=f"ААААА! {name} что ты делаешь?")
        update.message.reply_text(text="Я не понимаю")

def do_start(update, context):
    keyboard = [
        ["Худи", "Футболки", "Джинсы"],
        ["Шорты", "Ботинки", "назад"]
    ]
    update.message.reply_text(
        text="Привет, это подборка одежды",
        reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True))

def do_help(update, context):
    update.message.reply_text(text="что случилось? Всё ж было норм")

def do_sticker(update: Update, context):
    sticker_id = update.message.sticker.file_id
    update.message.reply_text(sticker_id)
    update.message.reply_sticker(sticker_id)

def do_something(update, context):
    text = update.message.text
    print(stickers_page.max_row)
    for row in range(2, stickers_page.max_row + 1):
        catch_phrase = stickers_page.cell(row=row, column=4).value
        print(catch_phrase)
        print(text)
        if catch_phrase in text:
            sticker_id = stickers_page.cell(row=row, column=3).value
            update.message.reply_text(sticker_id)
    if text == "Худи":
        update.message.reply_text(text="вы нажали Худи")
    elif text == "Футболки":
        update.message.reply_text(text="вы нажали Футболки")
    elif text == "Джинсы":
        update.message.reply_text(text="вы нажали Джинсы")
    elif text == "Шорты":
        update.message.reply_text(text="вы нажали Шорты")
    elif text == "Ботинки":
        update.message.reply_text(text="вы нажали Ботинки")
    elif text == "Назад":
        update.message.reply_text(text="вы нажали Назад")
    else:
        update.message.reply_text(text="вы нажали что-то еще")


main()